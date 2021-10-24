from openpyxl import load_workbook
wb=load_workbook(r"C:\Users\jimir\OneDrive - The American College of Greece\Desktop\ALL-matches.xlsx")

ws=wb['ALL']

teams={"100 Thieves":[1000,"LCS"],
"1907 Fenerbahce":[1000,"TCL"],
"5 Ronin":[1000,"TCL"],
"Afreeca Freecs":[1000,"LCK"],
"AGO ROGUE":[1000,"EU MASTERS"],
"ahq e-Sports Club":[1000,"PCS"],
"All Knights":[1000,"LLA"],
"Alpha Esports":[1000,"PCS"],
"Anorthosis Famagusta Esports":[1000,"EU MASTERS"],
"Astralis":[1000,"LEC"],
"Asus ROG Elite":[1000,"EU MASTERS"],
"Avant Gaming":[1000,"LCO"],
"AXIZ":[1000,"LJL"],
"Azules Esports":[1000,"LLA"],
"Berjaya Dragons":[1000,"PCS"],
"Berlin International Gaming":[1000,"EU MASTERS"],
"Besiktas Esports":[1000,"TCL"],
"Beyond Gaming":[1000,"PCS"],
"Bilibili Gaming":[1000,"LPL"],
"Black Star Gaming":[1000,"LCL"],
"BOOM Esports":[1000,"PCS"],
"BT Excel":[1000,"EU MASTERS"],
"Burning Core":[1000,"LJL"],
"Burst The Sky Esports":[1000,"VCS"],
"Cerberus Esports":[1000,"VCS"],
"Chief Esports Club":[1000,"LCO"],
"CLG":[1000,"LCS"],
"Cloud9":[1000,"LCS"],
"CR4ZY":[1000,"EU MASTERS"],
"Cream Real Betis":[1000,"EU MASTERS"],
"Crest Gaming Act":[1000,"LJL"],
"CrowCrowd":[1000,"LCL"],
"CTRL PLAY":[1000,"LCL"],
"Cyber Gaming":[1000,"EU MASTERS"],
"Dark Passage":[1000,"TCL"],
"Defusekids":[1000,"EU MASTERS"],
"Detonation FocusMe":[1000,"LJL"],
"Dignitas":[1000,"LCS"],
"Dire Wolves":[1000,"LCO"],
"Dominus Esports":[1000,"LPL"],
"Dragon Army":[1000,"LCL"],
"DRX":[1000,"LCK"],
"DWG KIA":[1000,"LCK"],
"Edward Gaming":[1000,"LPL"],
"EGN Esports":[1000,"EU MASTERS"],
"Elements Pro Gaming":[1000,"LCL"],
"eStar":[1000,"LPL"],
"Estral Esports":[1000,"LLA"],
"eSuba":[1000,"EU MASTERS"],
"Evil Geniuses":[1000,"LCS"],
"EVOS Esports":[1000,"VCS"],
"Excel Esports":[1000,"LEC"],
"FC Schalke 04 Evolution":[1000,"EU MASTERS"],
"FC Schalke 04":[1000,"LEC"],
"Five Kings":[1000,"EU MASTERS"],
"Flamengo eSports":[1000,"CBLOL"],
"FlyQuest":[1000,"LCS"],
"Fnatic Rising":[1000,"EU MASTERS"],
"Fnatic":[1000,"LEC"],
"For The Win Esports":[1000,"EU MASTERS"],
"For The Win":[1000,"EU MASTERS"],
"Fredit BRION":[1000,"LCK"],
"FTV Esports":[1000,"VCS"],
"Funplus Phoenix":[1000,"LPL"],
"FURIA Esports":[1000,"CBLOL"],
"FURIA Uppercut":[1000,"CBLOL"],
"Furious Gaming":[1000,"LLA"],
"G2 Arctic":[1000,"EU MASTERS"],
"G2 Esports":[1000,"LEC"],
"Galakticos":[1000,"TCL"],
"Galatasaray Esports":[1000,"TCL"],
"GAM Esports":[1000,"VCS"],
"Gambit Esports":[1000,"LCL"],
"GamerLegion":[1000,"EU MASTERS"],
"Gamers Origin":[1000,"EU MASTERS"],
"Gamespace Mediterranean College Esports":[1000,"EU MASTERS"],
"Gen.G eSports":[1000,"LCK"],
"GMedia Luxury":[1000,"VCS"],
"Golden Guardians":[1000,"LCS"],
"Goskilla":[1000,"EU MASTERS"],
"Gravitas":[1000,"LCO"],
"Griffin":[1000,"LCK"],
"Hanwha Life eSports":[1000,"LCK"],
"Hong Kong Attitude":[1000,"PCS"],
"Illuminar Gaming":[1000,"EU MASTERS"],
"Immortals":[1000,"LCS"],
"Impunity":[1000,"PCS"],
"Infinity eSports":[1000,"LLA"],
"Intreprid Fox Gaming":[1000,"EU MASTERS"],
"INTZ e-Sports":[1000,"CBLOL"],
"Invictus Gaming":[1000,"LPL"],
"Iron Wolves":[1000,"EU MASTERS"],
"Istanbul Wildcats":[1000,"TCL"],
"Isurus Gaming":[1000,"LLA"],
"Isurus":[1000,"LLA"],
"J Team":[1000,"PCS"],
"JD Gaming":[1000,"LPL"],
"K1CK Neosurf":[1000,"EU MASTERS"],
"K1CK":[1000,"EU MASTERS"],
"KaBuM! e-Sports":[1000,"CBLOL"],
"Kaos Latin Gamers":[1000,"LLA"],
"Karmine Corp":[1000,"EU MASTERS"],
"Kenty":[1000,"EU MASTERS"],
"KT Rolster":[1000,"LCK"],
"KV Mechelen Esports":[1000,"EU MASTERS"],
"LDLC OL":[1000,"EU MASTERS"],
"Legacy Esports":[1000,"LCO"],
"LGD Gaming":[1000,"LPL"],
"Liiv SANDBOX":[1000,"LCK"],
"Liyab Esports":[1000,"PCS"],
"LNG Esports":[1000,"LPL"],
"LOUD":[1000,"CBLOL"],
"LowLandLions":[1000,"EU MASTERS"],
"M19":[1000,"LCL"],
"Machi E-Sports":[1000,"PCS"],
"Macko Esports":[1000,"EU MASTERS"],
"MAD Lions Madrid":[1000,"EU MASTERS"],
"MAD Lions":[1000,"LEC"],
"Mammoth":[1000,"LCO"],
"Method2Madness":[1000,"EU MASTERS"],
"Misfits Gaming":[1000,"LEC"],
"Misfits Premier":[1000,"EU MASTERS"],
"Mkers":[1000,"EU MASTERS"],
"Mousesports":[1000,"EU MASTERS"],
"Movistar Riders":[1000,"EU MASTERS"],
"NASR eSports Turkey":[1000,"TCL"],
"Netshoes Miners":[1000,"CBLOL"],
"Nongshim RedForce":[1000,"LCK"],
"Nova Esports":[1000,"PCS"],
"OMG":[1000,"LPL"],
"One Breath Gaming":[1000,"LCL"],
"Order":[1000,"LCO"],
"OverPower Esports":[1000,"VCS"],
"paiN Gaming":[1000,"CBLOL"],
"Papara SuperMassive":[1000,"TCL"],
"PEACE":[1000,"LCO"],
"Pentanet.GG":[1000,"LCO"],
"Percent Esports":[1000,"VCS"],
"PIGSPORTS":[1000,"EU MASTERS"],
"Pixel Esports Club":[1000,"LLA"],
"Pompa Team":[1000,"EU MASTERS"],
"Prodigy Esports":[1000,"CBLOL"],
"PSG Talon":[1000,"PCS"],
"PSV Esports":[1000,"EU MASTERS"],
"Racoon":[1000,"EU MASTERS"],
"Rainbow7":[1000,"LLA"],
"Rare Atom":[1000,"LPL"],
"Rascal Jester":[1000,"LJL"],
"RED Canids":[1000,"CBLOL"],
"Redemption POA":[1000,"CBLOL"],
"Rensga eSports":[1000,"CBLOL"],
"Resurgence":[1000,"PCS"],
"Riddle Esports":[1000,"EU MASTERS"],
"Rogue Warriors":[1000,"LPL"],
"Rogue":[1000,"LEC"],
"RoX":[1000,"LCL"],
"Royal Never Give Up":[1000,"LPL"],
"Royal Youth":[1000,"TCL"],
"Saigon Buffalo":[1000,"VCS"],
"SAIM SE SuppUp":[1000,"EU MASTERS"],
"SAMCLAN Esports":[1000,"EU MASTERS"],
"Samsung Morning Stars":[1000,"EU MASTERS"],
"Sandbox Gaming":[1000,"LCK"],
"Santos e-Sports":[1000,"CBLOL"],
"SBTC Esports":[1000,"VCS"],
"Schalke Evolution":[1000,"EU MASTERS"],
"Sector One":[1000,"EU MASTERS"],
"Sengoku Gaming":[1000,"LJL"],
"SeolHaeOne Prince":[1000,"LCK"],
"SINNERS Esports":[1000,"EU MASTERS"],
"SK Gaming":[1000,"LEC"],
"SoftBank Hawks Gaming":[1000,"LJL"],
"Suning":[1000,"LPL"],
"SuperMassive Blaze":[1000,"TCL"],
"SuppUp eSports":[1000,"EU MASTERS"],
"T1":[1000,"LCK"],
"Team Aurora":[1000,"TCL"],
"Team Dynamics":[1000,"LCK"],
"Team Flash":[1000,"VCS"],
"Team Liquid":[1000,"LCS"],
"Team Secret":[1000,"VCS"],
"Team Singularity":[1000,"EU MASTERS"],
"Team Vitality":[1000,"LEC"],
"Team WE":[1000,"LPL"],
"Top Esports":[1000,"LPL"],
"TSM":[1000,"LCS"],
"TT":[1000,"LPL"],
"UCAM Esports Club":[1000,"EU MASTERS"],
"Ultra Prime":[1000,"LPL"],
"Unicorns Of Love":[1000,"LCL"],
"V Gaming Adonis":[1000,"VCS"],
"V3 Esports":[1000,"LJL"],
"Vega Squadron":[1000,"LCL"],
"Vici Gaming":[1000,"LPL"],
"Victory Five":[1000,"LPL"],
"Vipers Inc":[1000,"EU MASTERS"],
"Vitality.Bee":[1000,"EU MASTERS"],
"Vivo Keyd":[1000,"CBLOL"],
"Vodafone Giants":[1000,"EU MASTERS"],
"Vorax Liberty":[1000,"CBLOL"],
"White Dragons":[1000,"EU MASTERS"],
"WLGaming Esports":[1000,"EU MASTERS"],
"XTEN Esports":[1000,"LLA"],
"YDN Gamers":[1000,"EU MASTERS"],
"Zero Tenacity":[1000,"EU MASTERS"],
"Karma Clan Esports":[1000,"EU MASTERS"],
"Team Phantasma":[1000,"EU MASTERS"],
"Crvena zvezda Esports":[1000,"EU MASTERS"],
"Outplayed":[1000,"EU MASTERS"],
"mCon esports Rotterdam":[1000,"EU MASTERS"],
"Vodafone Giants.Spain":[1000,"EU MASTERS"],
"Tricked Esports":[1000,"EU MASTERS"],
"PDW":[1000,"EU MASTERS"],
"PENTA 1860":[1000,"EU MASTERS"],
"GOEXANIMO":[1000,"EU MASTERS"]
}

regions={"EU MASTERS":[0,0],
         "LPL":[0,0],
         "LEC":[0,0],
         "LCK":[0,0],
         "LCS":[0,0],
         "CBLOL":[0,0],
         "VCS":[0,0],
         "PCS":[0,0],
         "TCL":[0,0],
         "LJL":[0,0],
         "LLA":[0,0],
         "LCL":[0,0],
         "LCO":[0,0],
         }

#fix ratings to regions

for i in teams:
    if teams[i][1]=="LPL":
        teams[i][0]+=140
    elif teams[i][1]=="LCK":
        teams[i][0]+=150
    elif teams[i][1]=="EU MASTERS":
        teams[i][0]+=-400
    elif teams[i][1]=="LEC":
        teams[i][0]+=50
    elif teams[i][1]=="CBLOL":
        teams[i][0]+=-300
    elif teams[i][1]=="VCS":
        teams[i][0]+=-300
    elif teams[i][1]=="PCS":
        teams[i][0]+=-350
    elif teams[i][1]=="TCL":
        teams[i][0]+=-220
    elif teams[i][1]=="LJL":
        teams[i][0]+=-200
    elif teams[i][1]=="LLA":
        teams[i][0]+=-300
    elif teams[i][1]=="LCL":
        teams[i][0]+=-350
    elif teams[i][1]=="LC0":
        teams[i][0]+=-150
   

    
probs={800:[0,0],750:[0,0],700:[0,0],650:[0,0],600:[0,0],550:[0,0],500:[0,0],
      450:[0,0],375:[0,0],325:[0,0],275:[0,0],225:[0,0],175:[0,0],
      125:[0,0],75:[0,0],25:[0,0],0:[0,0]}      

h=2
k=37
d=160
for i in range(2,h):

    e1=10**(teams[ws.cell(row=i,column=2).value][0]/d)
    e2=10**(teams[ws.cell(row=i,column=5).value][0]/d)
    r1=int(ws.cell(row=i,column=3).value)
    r2=int(ws.cell(row=i,column=4).value)
    s1=r1/(r1+r2)
    s2=r2/(r1+r2)
        

    #calculation

    teams[ws.cell(row=i,column=2).value][0]+=k*(s1-(e1/(e1+e2)))
    teams[ws.cell(row=i,column=5).value][0]+=k*(s2-(e2/(e1+e2)))

for i in range(h,4936):     
    e1=10**(teams[ws.cell(row=i,column=2).value][0]/d)
    e2=10**(teams[ws.cell(row=i,column=5).value][0]/d)
    r1=int(ws.cell(row=i,column=3).value)
    r2=int(ws.cell(row=i,column=4).value)
    s1=r1/(r1+r2)
    s2=r2/(r1+r2)

    #regions

    if teams[ws.cell(row=i,column=2).value][1]!=teams[ws.cell(row=i,column=5).value][1]:
        if ws.cell(row=i,column=3).value>ws.cell(row=i,column=4).value:
            regions[teams[ws.cell(row=i,column=2).value][1]][0]+=1
        else:
            regions[teams[ws.cell(row=i,column=2).value][1]][1]+=1

    #prediction
    dif=abs(teams[ws.cell(row=i,column=2).value][0]-teams[ws.cell(row=i,column=5).value][0])
    if (e1>e2 and s1>s2) or (e2>e1 and s2>s1):       
        for j in probs:
                if dif>j:
                    probs[j][0]+=1
                    break
    elif (e1>e2 and s1<s2)or(e1<e2 and s1>s2):
        for j in probs:
                if dif>j:
                    probs[j][1]+=1
                    #print("lost match ",ws.cell(row=i,column=2).value," vs ",ws.cell(row=i,column=5).value)
                    break
 
    #calculation

    teams[ws.cell(row=i,column=2).value][0]+=k*(s1-(e1/(e1+e2)))
    teams[ws.cell(row=i,column=5).value][0]+=k*(s2-(e2/(e1+e2)))   
    


sort_teams=sorted(teams.items(),key=lambda x:x[1], reverse=True)


majorregions={"LPL","LEC","LCS","LCK"}
for i in sort_teams:
   # if i[1][1] in majorregions:
        print(i[0],":",i[1])

for i in probs:
    if (probs[i][0]+probs[i][1])!=0:
        print(i," : ",probs[i], " : ",probs[i][0]/(probs[i][0]+probs[i][1]))

for i in regions:
    print(i," : ", regions[i][0],"-",regions[i][1])
